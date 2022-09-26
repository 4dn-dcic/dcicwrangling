from scripts import geo2fdn as geo
from functions.notebook_functions import digest_xlsx
import GEOparse
import openpyxl
import pytest
import os


@pytest.fixture
def srx_file():
    return './tests/data_files/SRX3028942.xml'


@pytest.fixture
def bs_obj(mocker):
    with open('./tests/data_files/SAMN06219555.xml', 'r') as samn:
        mocker.patch('Bio.Entrez.efetch', return_value=samn)
        return geo.parse_bs_record('SAMN06219555')


@pytest.fixture
def exp_with_sra(mocker, srx_file):
    with open(srx_file, 'r') as srx:
        mocker.patch('Bio.Entrez.efetch', return_value=srx)
        gsm = GEOparse.get_GEO(filepath='./tests/data_files/GSM2715320.txt')
        return geo.parse_gsm(gsm)


@pytest.fixture
def exp_with_sra_pe(mocker):
    with open('./tests/data_files/SRX1839065.xml', 'r') as srx:
        mocker.patch('Bio.Entrez.efetch', return_value=srx)
        gsm = GEOparse.get_GEO(filepath='./tests/data_files/GSM2198225.txt')
        return geo.parse_gsm(gsm)


@pytest.fixture
def hidden_sra(mocker):
    with open('./tests/data_files/SRX4191023.xml', 'r') as srx:
        mocker.patch('Bio.Entrez.efetch', return_value=srx)
        gsm = GEOparse.get_GEO(filepath='./tests/data_files/GSM2715320.txt')
        return geo.parse_gsm(gsm)


def test_parse_gsm_with_sra(mocker, srx_file):
    with open(srx_file, 'r') as srx:
        mocker.patch('Bio.Entrez.efetch', return_value=srx)
        gsm = GEOparse.get_GEO(filepath='./tests/data_files/GSM2715320.txt')
        exp = geo.parse_gsm(gsm)
    assert exp.link in srx_file
    assert exp.exptype == 'repliseq'
    assert exp.bs == 'SAMN06219555'
    assert exp.layout == 'single'
    assert exp.instr == 'Ion Torrent Proton'
    assert len(exp.runs) == 1
    assert exp.length == 51


def test_parse_gsm_dbgap(mocker):
    with open('./tests/data_files/SRX3028942.xml', 'r') as srx:
        mocker.patch('Bio.Entrez.efetch', return_value=srx)
        gsm = GEOparse.get_GEO(filepath='./tests/data_files/GSM2254215.txt')
        exp = geo.parse_gsm(gsm)
    assert exp.bs == 'SAMN05449633'
    assert not exp.layout
    assert exp.instr == 'Illumina HiSeq 2500'
    assert exp.exptype.startswith('hic')
    assert not exp.link
    assert not exp.runs
    assert not exp.length


def test_experiment_bad_sra(mocker, capfd):
    with open('./tests/data_files/SRX0000000.xml', 'r') as srx:
        mocker.patch('Bio.Entrez.efetch', return_value=srx)
        exp = geo.Experiment('hic', 'Illumina HiSeq 2500', 'GSM1234567', 'title', 'SAMN05449633', 'SRX0000000')
        exp.get_sra()
    out, err = capfd.readouterr()
    assert "Couldn't parse" in out


def gsm_soft_to_exp_obj(mocker, gsm_file, exp_type=None):
    mocker.patch('scripts.geo2fdn.Experiment.get_sra')
    gsm = GEOparse.get_GEO(filepath=gsm_file)
    return geo.parse_gsm(gsm, experiment_type=exp_type)


def test_parse_gsm_exptypes(mocker):
    soft_file_dict = {'GSM3154187': 'dna sprite', 'GSM2198225': 'capturec',
                      'GSM3149191': 'atacseq', 'GSM2586973': 'damidseq', 'GSM3003988': 'chiapet'}
    for acc in soft_file_dict.keys():
        parsed = gsm_soft_to_exp_obj(mocker, './tests/data_files/' + acc + '.txt')
        assert parsed.exptype == soft_file_dict[acc]


def test_parse_bs_record(mocker):
    with open('./tests/data_files/SAMN06219555.xml', 'r') as samn:
        mocker.patch('Bio.Entrez.efetch', return_value=samn)
        bs = geo.parse_bs_record('SAMN06219555')
    for item in ['tamoxifen', 'liver', 'NIPBL', 'Nipbl(flox/flox)']:
        assert item in bs.description


def test_get_geo_metadata_seq(mocker):
    mocker.patch('scripts.geo2fdn.Experiment.get_sra')
    mocker.patch('scripts.geo2fdn.parse_bs_record', return_value='SAMNXXXXXXXX')
    gse = geo.get_geo_metadata('./tests/data_files/GSE93431_family.soft.gz')
    assert len([exp for exp in gse.experiments if exp.exptype == 'hic']) == 6
    assert len([exp for exp in gse.experiments if exp.exptype == 'chipseq']) == 14
    assert len([exp for exp in gse.experiments if exp.exptype == 'rnaseq']) == 12
    assert len([bs for bs in gse.biosamples]) == 32


def test_get_geo_metadata_microarray(capfd):
    gse = geo.get_geo_metadata('./tests/data_files/GSE102960_family.soft.gz')
    out, err = capfd.readouterr()
    assert not gse
    assert out == 'Sequencing experiments not found. Exiting.\n'


def test_get_geo_metadata_bad_accession(capfd):
    gse = geo.get_geo_metadata('GDS102960')
    out, err = capfd.readouterr()
    assert not gse
    assert out == 'Input not a valid GEO accession.\n'


def test_get_geo_metadata_sra_hidden(capfd, mocker, hidden_sra):
    gse_all = GEOparse.get_GEO(filepath='./tests/data_files/GSE93431_family.soft.gz')
    mocker.patch('scripts.geo2fdn.parse_bs_record', return_value='SAMNXXXXXXXX')
    mocker.patch('scripts.geo2fdn.parse_gsm', return_value=hidden_sra)
    gse = geo.get_geo_metadata('./tests/data_files/GSE93431_family.soft.gz')
    out, err = capfd.readouterr()
    assert not gse
    assert len(gse_all.gsms.values()) > 10


def create_xlsx_dict(inbook, sheetnames):
    xlsx_dict = {}
    for sheetname in sheetnames:
        current_sheet = inbook[sheetname]
        for row_index, row in enumerate(current_sheet.values, start=1):
            if row_index == 5:
                break
        if row_index > 4:
            col_dict = {}
            for col in current_sheet.iter_cols(values_only=True):
                header = col[0]
                col_dict[header] = [value for value in col[4:]]
            xlsx_dict[sheetname] = col_dict
    return xlsx_dict


def test_get_geo_tables(mocker, bs_obj, exp_with_sra):
    mocker.patch('scripts.geo2fdn.parse_gsm', return_value=exp_with_sra)
    mocker.patch('scripts.geo2fdn.parse_bs_record', return_value=bs_obj)
    geo.get_geo_tables('./tests/data_files/GSM2715320.txt', './tests/data_files/test_table', email='test@email.com')
    with open('./tests/data_files/test_table_fqs.tsv', 'r') as fq_file:
        fq_out = fq_file.readlines()
    assert not any('paired with' in line for line in fq_out)
    for item in ['expts', 'fqs', 'bs']:
        os.remove(f'./tests/data_files/test_table_{item}.tsv')


def test_get_geo_tables_pe(mocker, bs_obj, exp_with_sra_pe):
    mocker.patch('scripts.geo2fdn.parse_gsm', return_value=exp_with_sra_pe)
    mocker.patch('scripts.geo2fdn.parse_bs_record', return_value=bs_obj)
    geo.get_geo_tables('./tests/data_files/GSM2198225.txt', './tests/data_files/test_table', email='test@email.com')
    with open('./tests/data_files/test_table_fqs.tsv', 'r') as fq_file:
        fq_out = fq_file.readlines()
    assert all('paired with' in line for line in fq_out)
    for item in ['expts', 'fqs', 'bs']:
        os.remove(f'./tests/data_files/test_table_{item}.tsv')


def test_modify_xlsx(mocker, bs_obj, exp_with_sra):
    mocker.patch('scripts.geo2fdn.parse_gsm', return_value=exp_with_sra)
    mocker.patch('scripts.geo2fdn.parse_bs_record', return_value=bs_obj)
    # gds = geo.get_geo_metadata('GSM2715320', filepath='./tests/data_files/GSM2715320.txt')
    geo.modify_xlsx('./tests/data_files/GSM2715320.txt', './tests/data_files/repliseq_template.xlsx', 'out.xlsx', 'abc')
    book, sheetnames = digest_xlsx('out.xlsx')
    outfile_dict = create_xlsx_dict(book, sheetnames)
    os.remove('out.xlsx')
    assert outfile_dict['Biosample']['aliases'][0].startswith('abc:')
    assert outfile_dict['Biosample']['dbxrefs'][0].startswith('BioSample:SAMN')
    # assert BiosampleCellCulture has alias
    assert (outfile_dict['BiosampleCellCulture']['aliases'][0].startswith('abc:')
            and outfile_dict['BiosampleCellCulture']['aliases'][0].endswith('-cellculture'))
    # assert BiosampleCellCulture alias is in Biosample sheet
    assert (outfile_dict['Biosample']['cell_culture_details'][0].startswith('abc:')
            and outfile_dict['Biosample']['cell_culture_details'][0].endswith('-cellculture'))
    # FileFastq assert(s)
    assert outfile_dict['FileFastq']['*file_format'][0] == 'fastq'
    assert not outfile_dict['FileFastq']['paired_end'][0]
    assert not (outfile_dict['FileFastq']['related_files.relationship_type'][0]
                or outfile_dict['FileFastq']['related_files.file'][0])
    assert outfile_dict['FileFastq']['read_length'][0]
    assert outfile_dict['FileFastq']['instrument'][0]
    assert outfile_dict['FileFastq']['dbxrefs'][0].startswith('SRA:SRR')
    # ExperimentRepliseq assert(s)
    assert outfile_dict['ExperimentRepliseq']['dbxrefs'][0].startswith('GEO:GSM')
    assert outfile_dict['ExperimentRepliseq']['description'][0]
    assert outfile_dict['ExperimentRepliseq']['files'][0]
    assert outfile_dict['ExperimentRepliseq']['*biosample'][0]


def test_modify_xlsx_pe(mocker, bs_obj, exp_with_sra_pe):
    mocker.patch('scripts.geo2fdn.parse_gsm', return_value=exp_with_sra_pe)
    mocker.patch('scripts.geo2fdn.parse_bs_record', return_value=bs_obj)
    geo.modify_xlsx('./tests/data_files/GSM2198225.txt',
                    './tests/data_files/capturec_seq_template.xlsx',
                    'out_pe.xlsx', 'abc')
    book, sheetnames = digest_xlsx('out_pe.xlsx')
    outfile_dict = create_xlsx_dict(book, sheetnames)
    os.remove('out_pe.xlsx')
    # FileFastq assert(s)
    assert outfile_dict['FileFastq']['*file_format'][0] == 'fastq'
    assert outfile_dict['FileFastq']['paired_end'][0]
    assert (outfile_dict['FileFastq']['related_files.relationship_type'][0]
            and outfile_dict['FileFastq']['related_files.file'][0])
    assert outfile_dict['FileFastq']['read_length'][0]
    assert outfile_dict['FileFastq']['instrument'][0]
    assert outfile_dict['FileFastq']['dbxrefs'][0].startswith('SRA:SRR')


def test_modify_xlsx_some_unparsable_types(mocker, capfd, bs_obj):
    mocker.patch('scripts.geo2fdn.Experiment.get_sra')
    mocker.patch('scripts.geo2fdn.parse_bs_record', return_value=bs_obj)
    geo.modify_xlsx('./tests/data_files/GSE87585_family.soft.gz',
                    './tests/data_files/capturec_seq_template.xlsx', 'out2.xlsx', 'abc')
    out, err = capfd.readouterr()
    book, sheetnames = digest_xlsx('out2.xlsx')
    outfile_dict = create_xlsx_dict(book, sheetnames)
    os.remove('out2.xlsx')
    assert len(outfile_dict['ExperimentSeq']['aliases']) > 0
    types_in_outfile = outfile_dict['ExperimentSeq']['*experiment_type']
    lines = out.split('\n')
    assert 'RNA-seq' in types_in_outfile and len(types_in_outfile) == 6
    assert 'ExperimentCaptureC' not in outfile_dict.keys()
    assert 'The following accessions had experiment types that could not be parsed:' in lines
    assert 'HiC experiments found in ./tests/data_files/GSE87585_family.soft.gz but no ExperimentHiC sheet' in lines


def test_modify_xlsx_set_experiment_type(mocker, capfd, bs_obj):
    mocker.patch('scripts.geo2fdn.Experiment.get_sra')
    mocker.patch('scripts.geo2fdn.parse_bs_record', return_value=bs_obj)
    geo.modify_xlsx('./tests/data_files/GSE87585_family.soft.gz',
                    './tests/data_files/capturec_seq_template.xlsx', 'out3.xlsx', 'abc',
                    experiment_type='CaptureC')
    book, sheetnames = digest_xlsx('out3.xlsx')
    out, err = capfd.readouterr()
    outfile_dict = create_xlsx_dict(book, sheetnames)
    os.remove('out3.xlsx')
    assert 'ExperimentSeq' not in outfile_dict.keys()
    assert len(outfile_dict['ExperimentCaptureC']['aliases']) == 12
    assert 'The following accessions had experiment types that could not be parsed:' not in out.split('\n')


def test_modify_xlsx_no_sheets(mocker, bs_obj):
    mocker.patch('scripts.geo2fdn.Experiment.get_sra')
    mocker.patch('scripts.geo2fdn.parse_bs_record', return_value=bs_obj)
    geo.modify_xlsx('./tests/data_files/GSE87585_family.soft.gz',
                    './tests/data_files/no_template.xlsx', 'out4.xlsx', 'abc',
                    experiment_type='CaptureC')
    book, sheetnames = digest_xlsx('out4.xlsx')
    outfile_dict = create_xlsx_dict(book, sheetnames)
    os.remove('out4.xlsx')
    assert not outfile_dict


def run_compare(capfd, exp_with_sra, template, exp_type, sheet):
    inbook = digest_xlsx(template)[0]
    exp_list = [exp for exp in [exp_with_sra] if exp.exptype == exp_type]
    acc = exp_with_sra.geo
    geo.experiment_type_compare(sheet, exp_list, acc, inbook)
    out, err = capfd.readouterr()
    return out.split('\n'), acc


def test_experiment_type_compare_nosheet_exp(capfd, exp_with_sra):
    out, acc = run_compare(capfd, exp_with_sra, './tests/data_files/capturec_seq_template.xlsx',
                           'repliseq', 'ExperimentRepliseq')
    assert f'Repliseq experiments found in {acc} but no ExperimentRepliseq sheet' in out


def test_experiment_type_compare_sheet_noexp(capfd, exp_with_sra):
    out, acc = run_compare(capfd, exp_with_sra, './tests/data_files/capturec_seq_template.xlsx',
                           'capturec', 'ExperimentCaptureC')
    assert f'No CaptureC experiments parsed from {acc}.' in out


def test_experiment_type_compare_sheet_exp(capfd, exp_with_sra):
    out, acc = run_compare(capfd, exp_with_sra, './tests/data_files/repliseq_template.xlsx',
                           'repliseq', 'ExperimentRepliseq')
    assert f'No Repliseq experiments parsed from {acc}.' not in out
    assert f'Repliseq experiments found in {acc} but no ExperimentRepliseq sheet' not in out
