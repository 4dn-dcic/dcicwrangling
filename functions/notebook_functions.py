from dcicutils import ff_utils
from uuid import UUID
import os
import json
import openpyxl
import warnings  # to suppress openpxl warning about headers
from openpyxl.utils.exceptions import InvalidFileException
import datetime


def digest_xlsx(filename):
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            book = openpyxl.load_workbook(filename)
    except InvalidFileException as e:
        if filename.endswith('.xls'):
            print("WARNING - Old xls format not supported - please save your workbook as xlsx")
        else:
            print("ERROR - ", e)
        sys.exit(1)
    sheets = book.sheetnames
    return book, sheets


def reader(workbook, sheetname=None):
    """Read named sheet or first and only sheet from xlsx file."""
    if sheetname is None:
        sheet = workbook.worksheets[0]
    else:
        try:
            sheet = workbook[sheetname]
        except Exception as e:
            print(e)
            print(sheetname)
            print("ERROR: Can not find the collection sheet in excel file (openpyxl error)")
            return
    # Generator that gets rows from excel sheet
    # NB we have a lot of empty no formatting rows added (can we get rid of that)
    # or do we need to be careful to check for the first totally emptyvalue row?
    return row_generator(sheet)


def row_generator(sheet):
    """Generator that gets rows from excel sheet
    Note that this currently checks to see if a row is empty and if so stops
    This is needed as plain text formatting of cells is recognized as data
    """
    for row in sheet.rows:
        vals = [cell_value(cell) for cell in row]
        if not any([v for v in vals]):
            return
        else:
            yield vals


def cell_value(cell):
    """Get cell value from excel. [From Submit4DN]"""
    ctype = cell.data_type
    value = cell.value
    if ctype == openpyxl.cell.cell.TYPE_ERROR:  # pragma: no cover
        raise ValueError('Cell %s contains a cell error' % str(cell.coordinate))
    elif ctype == openpyxl.cell.cell.TYPE_BOOL:
        boolstr = str(value).strip()
        if boolstr == 'TRUE':
            return True
        elif boolstr == 'FALSE':
            return False
        else:
            return value
    elif ctype in (openpyxl.cell.cell.TYPE_NUMERIC, openpyxl.cell.cell.TYPE_NULL):
        if isinstance(value, float):
            if value.is_integer():
                value = int(value)
        if not value:
            return ''
        return value
    elif isinstance(value, openpyxl.cell.cell.TIME_TYPES):
        if isinstance(value, datetime.datetime):
            if value.time() == datetime.time(0, 0, 0):
                return value.date().isoformat()
            else:  # pragma: no cover
                return value.isoformat()
        else:
            return value.isoformat()
    elif ctype in (openpyxl.cell.cell.TYPE_STRING, openpyxl.cell.cell.TYPE_INLINE):
        return value.strip()
    raise ValueError(
        'Cell %s is not an acceptable cell type' % str(cell.coordinate)
    )  # pragma: no cover


def get_key(keyname=None, keyfile='keypairs.json'):
    """get the key from your keypairs.json file
    if no keyname is given, use default key,
    but ask before moving on, if keyfile is given,
    keyname is a must"""
    # is there a different keyfile?
    if keyfile != 'keypairs.json':
        if not keyname:
            raise Exception('please provide keyname')
        if keyfile.startswith('~'):
            keyfile = os.path.expanduser("~") + keyfile[1:]
        with open(keyfile, 'r') as key_file:
            keys = key_file.read()
        my_key = json.loads(keys)[keyname]
    else:
        home_dir = os.path.expanduser("~") + "/"
        key_file = home_dir + keyfile
        keys = open(key_file, 'r').read()
        if not keyname:
            my_key = json.loads(keys)['default']
            print(my_key['server'])
            go_on = input("Using the 'default' key? (Y/N)")
            if go_on.lower() in ['y', 'yes']:
                pass
            else:
                raise Exception('please provide your keyname parameter')
        else:
            my_key = json.loads(keys)[keyname]
    return my_key


def list_column_titles(sheet):
    '''
        Return a list of values in the first row of an xlsx sheet (typically column titles)
    '''
    col_titles = [cell for (cell,) in sheet.iter_cols(min_row=1, max_row=1, values_only=True)]
    return col_titles


def copy_xlsx_sheet(sheet_to_copy, workbook, sheet_title):
    '''
        Create a new sheet in the target workbook with the input title and
        copy all values from the input sheet.
    '''
    new_sheet = workbook.create_sheet(title=sheet_title)
    for row_index, row in enumerate(sheet_to_copy.values, start=1):
        for col_index, cell_value in enumerate(row, start=1):
            new_sheet.cell(row=row_index, column=col_index, value=cell_value)
    return new_sheet, row_index


def append_xlsx_rows_formatted(sheet, formatted_items, row_index):
    '''
        Given a sheet in an xlsx workbook and a list of items formatted to fit
        exactly this sheet, it appends all items as new rows to the sheet,
        starting AFTER the row index provided.
    '''
    for row_index_append, row_item in enumerate(formatted_items, start=row_index + 1):
        for col_index, cell_value in enumerate(row_item, start=1):
            sheet.cell(row=row_index_append, column=col_index, value=cell_value)
    return sheet


def append_items_to_xlsx(input_xlsx, add_items, schema_names, comment=True):
    output_file_name = "_with_items.".join(input_xlsx.split('.'))
    bookread = openpyxl.load_workbook(input_xlsx)
    book_w = openpyxl.Workbook()
    book_w.remove(book_w.active)  # removes the empty sheet created by default named Sheet

    for sheet in bookread.sheetnames:
        active_sheet = bookread[sheet]
        first_row_values = list_column_titles(active_sheet)

        # create a new sheet and write the data
        new_sheet, n_copied_rows = copy_xlsx_sheet(active_sheet, book_w, sheet)

        # get items to add
        items_to_add = add_items.get(schema_names[sheet])
        if items_to_add:
            # exception for imaging paths
            if sheet == 'ExperimentMic':
                new_sheet, first_row_values = add_extra_path_columns(new_sheet, first_row_values, items_to_add)
            # append rows at the bottom
            formatted_items = format_items(items_to_add, first_row_values, comment)
            new_sheet = append_xlsx_rows_formatted(new_sheet, formatted_items, n_copied_rows)

    book_w.save(output_file_name)
    print('new excel is stored as', output_file_name)
    return


def add_extra_path_columns(new_sheet, first_row_values, items_to_add):
    '''
        Duplicates the imaging_paths.channel and imaging_paths.path columns
        in new_sheet if there is more than one path in items_to_add
    '''
    # find the longest number of imaging_paths in items_to_add
    img_path_keys = []
    for item in items_to_add:
        if item.get('imaging_paths'):
            current_img_path_keys = list(items_to_add[0]['imaging_paths'][0].keys())
            if len(current_img_path_keys) > len(img_path_keys):
                img_path_keys = current_img_path_keys

    # check if imaging_paths is in first_row_values
    try:
        imgpth_channel_index = first_row_values.index('imaging_paths.channel')
        # assume imaging_paths.path is at index imgpth_channel_index + 1
    except ValueError:
        return new_sheet, first_row_values

    # insert missing columns
    insert_index = imgpth_channel_index + 2
    for img_path in img_path_keys:
        full_img_path = 'imaging_paths.' + img_path
        if full_img_path in first_row_values:
            insert_index = first_row_values.index(full_img_path) + 1
        else:
            first_row_values.insert(insert_index, full_img_path)
            new_sheet.insert_cols(insert_index + 1)  # workbooks count from 1, not 0
            if img_path.startswith('channel'):
                index_column_to_duplicate = imgpth_channel_index + 1
            else:
                index_column_to_duplicate = imgpth_channel_index + 2
            column_to_duplicate = [v for (v,) in new_sheet.iter_rows(min_col=index_column_to_duplicate,
                                                                     max_col=index_column_to_duplicate,
                                                                     values_only=True)]
            column_to_duplicate[0] = full_img_path  # replace title
            for row_index, cell_value in enumerate(column_to_duplicate, start=1):
                new_sheet.cell(row=row_index, column=insert_index + 1, value=cell_value)
            insert_index += 1

    return new_sheet, first_row_values


def format_items(items_list, field_list, comment):
    """For a given sheet, get all released items"""
    all_items = []
    # filter for fields that exist on the excel sheet
    for item in items_list:
        item_info = []
        for field in field_list:
            write_value = ''
            # required fields will have a star
            field = field.strip('*')
            # add # to skip existing items during submission
            if field == "#Field Name:":
                if comment:
                    item_info.append("#")
                else:
                    item_info.append("")
            # the attachment field returns a dictionary
            elif field == "attachment":
                item_info.append("")
            else:
                # add sub-embedded objects
                # 1) only add if the field is not enumerated
                # 2) only add the first item if there are multiple
                # if you want to add more, accumulate all key value pairs in a single dictionary
                # [{main.sub1:a, main.sub2:b ,main.sub1-1:c, main.sub2-1:d,}]
                # and prepare the excel with these fields
                if "." in field:

                    main_field, sub_field = field.split('.')
                    temp_value = item.get(main_field)
                    if temp_value:
                        write_value = temp_value[0].get(sub_field, '')

                # usual cases
                else:
                    write_value = item.get(field, '')

                # take care of empty lists
                if not write_value:
                    write_value = ''

                # check for linked items
                if isinstance(write_value, dict):
                    write_value = write_value.get('@id')

                # when writing values, check for the lists and turn them into string
                if isinstance(write_value, list):
                    # check for linked items
                    if isinstance(write_value[0], dict):
                        write_value = [i.get('@id') for i in write_value]
                    write_value = ','.join(write_value)
                item_info.append(write_value)
        all_items.append(item_info)
    return all_items


def is_uuid(value):
    # md5 qualifies as uuid, not strictly uuid4: modify
    if '-' not in value:
        return False
    try:
        UUID(value, version=4)
        return True
    except:  # noqa
        return False


def find_uuids(val):
    vals = []
    if not val:
        return []
    elif isinstance(val, str):
        if is_uuid(val):
            vals = [val]
        else:
            return []
    else:
        text = str(val)
        text_list = [i for i in text. split("'") if len(i) == 36]
        vals = [i for i in text_list if is_uuid(i)]
    return vals


def get_schema_names(con_key):
    schema_name = {}
    profiles = ff_utils.get_metadata('/profiles/', key=con_key, add_on='frame=raw')
    for key, value in profiles.items():
        try:
            schema_name[key] = value['id'].split('/')[-1][:-5]
        except:
            continue
    return schema_name


def get_schema_names_and_fields(con_key):
    '''Gets concrete item types from profiles and returns a dict of schema
    names, with properties and property type (including if array_linkTo)'''
    schemas = {}
    profiles = ff_utils.get_metadata('/profiles/', key=con_key, add_on='frame=raw')
    for item in profiles.values():
        if item['isAbstract'] is True:
            continue
        schema_name = item['id'].split('/')[-1][:-5]
        schemas[schema_name] = {}
        for field, content in item['properties'].items():
            field_type = content['type']
            if field_type == 'array' and content['items'].get('linkTo'):
                field_type = 'array_linkTo'
            schemas[schema_name][field] = field_type
    return schemas


def dump_results_to_json(store, folder):
    if not os.path.exists(folder):
        os.makedirs(folder)
    for a_type in store:
        filename = folder + '/' + a_type + '.json'
        with open(filename, 'w') as outfile:
            json.dump(store[a_type], outfile, indent=4)


def printTable(myDict, colList=None):
    """ Pretty print a list of dictionaries Author: Thierry Husson"""
    if not colList:
        colList = list(myDict[0].keys() if myDict else [])
    myList = [colList]  # 1st row = header
    for item in myDict:
        myList.append([str(item[col] or '') for col in colList])
    colSize = [max(map(len, col)) for col in zip(*myList)]
    formatStr = ' | '.join(["{{:<{}}}".format(i) for i in colSize])
    myList.insert(1, ['-' * i for i in colSize])  # Seperating line
    for item in myList:
        print(formatStr.format(*item))


def clean_for_reupload(file_acc, key, clean_release_dates=False, delete_runs=True):
    """Rare cases we want to reupload the file, and this needs some cleanupself.
    If you want to delete release dates too, set 'clean_release_dates' to True"""
    resp = ff_utils.get_metadata(file_acc, key=key)
    clean_fields = ['extra_files', 'md5sum', 'content_md5sum', 'file_size', 'filename', 'quality_metric']
    if clean_release_dates:
        clean_fields.extend(['public_release', 'project_release'])
    if delete_runs:
        runs = resp.get('workflow_run_inputs', [])
        if runs:
            for a_run in runs:
                ff_utils.patch_metadata({'status': 'deleted'}, obj_id=a_run['uuid'], key=key)
    if resp.get('quality_metric'):
        ff_utils.patch_metadata({'status': 'deleted'}, obj_id=resp['quality_metric']['uuid'], key=key)
    del_f = []
    for field in clean_fields:
        if field in resp:
            del_f.append(field)
    del_add_on = 'delete_fields=' + ','.join(del_f)
    ff_utils.patch_metadata({'status': 'uploading'}, obj_id=resp['uuid'], key=key, add_on=del_add_on)


def file_in_exp(a_file, experiments):
    """Takes a file (as provided by get_es_metadata) and checks whether its
    experiment (or source_experiment) is found in a list of experiment uuids.
    If multiple experiments are associated with a file, returns a value only
    when all of them are present or absent in the experiment list.
    """
    found = False
    if a_file.get('experiments'):
        file_exps = [e['uuid'] for e in a_file['experiments']]
    elif a_file.get('source_experiments'):
        file_exps = a_file['source_experiments']  # this is already a list of uuids
    else:  # a file is not linked to any experiment
        found = None
        return found

    times_found = len([exp for exp in file_exps if exp in experiments])
    if times_found == len(file_exps):  # all experiments of a_file are in the exp list provided
        found = True
    elif times_found > 0:  # some, but not all of the exps of a_file are in the exp list provided
        found = None
    return found


def validate_change_helper(key, value, verb, types, statuses, level_min, level_max, SCHEMAS, STATUS_LEVEL):
    assert verb in ['add', 'remove', 'patch'], f'verb {verb} unknown'
    if verb in ['add', 'remove']:
        assert isinstance(value, list), 'add/remove can only be used on list fields'
    assert isinstance(types, list), f'types {types} is not a list'
    if types:
        for t in types:
            assert t in SCHEMAS, f'type {t} unknown'
            assert key in SCHEMAS[t], f'{key} not found in {t} schema'
            if SCHEMAS[t][key] == 'array_linkTo' and verb in ['add', 'remove']:
                for v in value:
                    assert is_uuid(v), 'add/remove sub-embedded objects works only with uuid'
    assert isinstance(statuses, list), f'statuses {statuses} is not a list'
    for s in statuses:
        assert s in STATUS_LEVEL, f'status {s} unknown'

    type_msg = f'{types}' if types else f'all types with field {key}'
    status_msg = f'in {statuses}' if statuses else f'level between {level_min} and {level_max} (included)'
    print(f'\nChanging field: {key}', f'Value: {value}', f'Action: {verb}',
          f'Item types to change: {type_msg}',
          f'Patch only items with final status {status_msg}', sep='\n')

    output_values = (key, value, verb, types, statuses, level_min, level_max)
    return output_values


def change_additional_fields(patch_body, item, item_type, item_level, change_level, change_status, changes, SCHEMAS):
    '''function to patch other properties, based on item type and final status (after update)'''
    for change in changes:
        (key, new_values, verb, types_to_change, statuses_to_change, level_min, level_max) = change
        # type check
        if types_to_change and item_type not in types_to_change:
            continue

        # key check
        if key not in SCHEMAS[item_type]:
            # not all items have fields such as `tags` or `viewable_by`
            continue

        # status check
        if statuses_to_change:
            new_item_status = change_status if (item_level != 0 and item_level < change_level) else item['status']
            if new_item_status not in statuses_to_change:
                continue
        else:
            new_item_level = max(item_level, change_level) if item_level != 0 else 0
            if new_item_level < level_min or new_item_level > level_max:
                continue

        # action depends on verb
        if verb == 'patch':
            patch_body[key] = new_values
        else:
            if verb == 'remove' and item.get(key) is None:
                # nothing to remove
                continue

            old_values = item.get(key, [])
            # 'add' and 'remove' work only on lists
            assert (isinstance(old_values, list)), 'add/remove can only be used on list fields'

            if old_values and isinstance(old_values[0], dict):
                # list of embedded objects: reduce to list of uuids
                old_values = [i.get('uuid') for i in old_values]

            if verb == 'add':
                new_unique_values = [v for v in new_values if v not in old_values]
                if len(new_unique_values) > 0:
                    patch_body[key] = [i for i in old_values] + new_unique_values
            elif verb == 'remove':
                remaining_old_values = [i for i in old_values if i not in new_values]
                if remaining_old_values != old_values:
                    patch_body[key] = remaining_old_values

    return patch_body


# get order from loadxl.py in fourfront
ORDER = [
    'user',
    'award',
    'lab',
    'static_section',
    'higlass_view_config',
    'page',
    'ontology',
    'ontology_term',
    'file_format',
    'badge',
    'organism',
    'genomic_region',
    'gene',
    'bio_feature',
    'target',
    'imaging_path',
    'publication',
    'publication_tracking',
    'document',
    'image',
    'vendor',
    'construct',
    'modification',
    'protocol',
    'sop_map',
    'experiment_type',
    'biosample_cell_culture',
    'individual_human',
    'individual_mouse',
    'individual_fly',
    'individual_chicken',
    'biosource',
    'antibody',
    'enzyme',
    'treatment_rnai',
    'treatment_agent',
    'biosample',
    'quality_metric_fastqc',
    'quality_metric_bamqc',
    'quality_metric_pairsqc',
    'quality_metric_dedupqc_repliseq',
    'quality_metric_chipseq',
    'quality_metric_atacseq',
    'quality_metric_workflowrun'
    'microscope_setting_d1',
    'microscope_setting_d2',
    'microscope_setting_a1',
    'microscope_setting_a2',
    'file_fastq',
    'file_processed',
    'file_reference',
    'file_calibration',
    'file_microscopy',
    'file_set',
    'file_set_calibration',
    'file_set_microscope_qc',
    'file_vistrack',
    'experiment_hi_c',
    'experiment_capture_c',
    'experiment_repliseq',
    'experiment_atacseq',
    'experiment_chiapet',
    'experiment_damid',
    'experiment_seq',
    'experiment_tsaseq',
    'experiment_mic',
    'experiment_set',
    'experiment_set_replicate',
    'data_release_update',
    'software',
    'analysis_step',
    'workflow',
    'workflow_mapping',
    'workflow_run_sbg',
    'workflow_run_awsem'
]
