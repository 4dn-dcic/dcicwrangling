#!/usr/bin/env python3
# -*- coding: latin-1 -*-

import argparse
import os
import re
import requests
import sys
import time
from statistics import mean
import xml.etree.ElementTree as ET
import openpyxl
from functions.notebook_functions import list_column_titles, copy_xlsx_sheet
from Bio import Entrez
from urllib.error import HTTPError
import GEOparse


description = '''
Script for fetching metadata from GEO and inserting it into a Submit4dn metadata workbook.

Note1: Use of NCBI's Entrez querying system requires an email address.
There will be a prompt to enter an email address when this script is run.

Note2: Currently only works if each experiment in GEO has an SRA record. Will not
work properly for sequences in dbgap, but hope to implement this in the future.

'''


epilog = '''
Example usage:

$ python scripts/geo2fdn.py GSE93431 -i hic_rnaseq_workbook.xlsx -o GSE93431_metadata.xlsx

$ python scripts/geo2fdn.py GSE68992 -i hic_workbook.xlsx -o GSE68992_metadata.xlsx -t 'dnase hic'

'''


class Experiment:

    def __init__(self, exptype, instr, geo, title, biosample, link):
        self.exptype = exptype.lower()  # experiment type
        self.instr = instr[0] if len(instr) == 1 else instr  # sequencing instrument
        self.layout = ''  # single or paired end
        self.geo = geo  # geo accession starting with GSM
        self.title = title[0] if len(title) == 1 else title
        self.runs = []  # list of SRA accessions starting with SRR
        self.length = ''  # mean read length
        # self.study_title = study_title
        self.bs = biosample
        self.link = link
        self.public = True

    def get_sra(self):
        # look up SRA record to fill out more attributes
        print('Fetching SRA Record...')
        handle = handle_timeout(Entrez.efetch(db="sra", id=self.link))
        try:
            record = ET.fromstring(handle.readlines()[2])
        except Exception:
            try:
                handle = handle_timeout(Entrez.efetch(db="sra", id=self.link))
                text = handle.read()
                record = ET.fromstring(text[text.index('<EXPERIMENT_PACKAGE>'):text.index('</EXPERIMENT_PACKAGE_SET>')])
            except Exception:
                print("Couldn't parse {} from {}".format(self.link, self.geo))
                return
        if record.tag == 'ERROR':
            if 'is not public' in record.text:
                print(record.text + ' - will not be written to file!')
                self.public = False
            else:
                print('Could not get SRA record {}'.format(self.link))
            return
        lengths = [int(float(item.get('average'))) for item in record.iter('Read') if item.get('count') != '0']
        if lengths:
            self.length = int(mean(lengths))
        self.st = record.find('STUDY').find('DESCRIPTOR').find('STUDY_TITLE').text
        for item in record.find('EXPERIMENT').find('DESIGN').find('LIBRARY_DESCRIPTOR').find('LIBRARY_LAYOUT'):
            self.layout = item.tag.lower()
            break
        self.runs = [item.get('accession') for item in record.find('RUN_SET').findall('RUN')]


class Biosample:

    def __init__(self, acc, organism, description):
        self.acc = acc  # BioSample accession starting with SAMN
        self.organism = organism
        self.description = description
        # self.treatments = treatments


class Dataset:

    def __init__(self, gse, ids, experiments, biosamples):
        self.gse = gse  # GEO Series accession starting with GSE
        self.ids = ids  # GEO sample ids associated with series
        self.experiments = experiments  # list of experiment objects
        self.biosamples = biosamples  # list of biosample objects


valid_types = ['hic', 'hicseq', 'dnase hic', 'rnaseq', 'tsaseq', 'chipseq',
               'dna sprite', 'dnarna sprite', 'rnadna sprite', 'capturec',
               'repliseq', 'atacseq', 'damid', 'damidseq', 'chiapet', 'placseq']


type_dict = {'chipseq': 'ChIP-seq', 'tsaseq': 'TSA-seq', 'rnaseq': 'RNA-seq',
             'atacseq': 'ATAC-seq', 'capturec': 'capture Hi-C', 'damid': 'DAM-ID seq',
             'damidseq': 'DAM-ID seq', 'chiapet': 'ChIA-PET', 'placseq': 'PLAC-seq',
             'dnase hic': 'DNase Hi-C', 'dna sprite': 'DNA SPRITE',
             'dnarna sprite': 'RNA-DNA SPRITE', 'rnadna sprite': 'RNA-DNA SPRITE',
             'repliseq': 'Repli-seq'}


def handle_timeout(command):  # pragma: no cover
    '''
    To retry commands if the server connection times out.
    '''
    time.sleep(0.5)
    try:
        result = command
    except HTTPError:
        time.sleep(3)
        try:
            result = command
        except HTTPError:
            time.sleep(10)
            try:
                result = command
            except HTTPError:
                time.sleep(10)
                result = command
    return result


def parse_gsm(gsm, experiment_type=None):
    '''
    Parses information about individual experiment. Input is a GEOparse.gsm object.
    Function creates an Experiment object; if GSM record has an associated SRA record,
    it will also look up the SRA record and fill out more attributes.
    '''
    # if experiment is a microarray, don't parse
    if 'SRA' not in gsm.metadata['type']:
        return
    exp_type = experiment_type if experiment_type else gsm.metadata['library_strategy'][0]
    if not exp_type or exp_type.lower() == 'other':
        for item in gsm.metadata['data_processing']:
            if item.lower().startswith('library strategy'):
                exp_type = item[item.index(':') + 2:]
    if not exp_type or exp_type.lower() == 'other':
        title = re.sub('-', '', ', '.join(gsm.metadata['title']).lower())
        types = [item for item in valid_types if item in title]
        if len(types) == 1 or (types and all('hic' in t for t in types)):
            exp_type = types[-1]
    link = None
    for item in gsm.metadata['relation']:
        # get biosample relation
        if item.startswith('BioSample:'):
            bs = item[item.index('SAMN'):]
        # get SRA relation
        elif item.startswith('SRA:'):
            link = item[item.index('SRX'):]
    exp = Experiment(re.sub('-', '', exp_type.lower()), gsm.metadata['instrument_model'],
                     gsm.name, gsm.metadata['title'], bs, link)
    if link:  # if no SRA relation is in GSM metadata, sequencing data might be in dbgap
        exp.get_sra()  # get more metadata about sequencing runs
    return exp


def get_geo_metadata(acc, experiment_type=None):
    '''
    Parses information associated with a GEO Series or single experiment.
    Uses GEOparse library which downloads records from NCBI ftp rather than using
    NCBI Entrez e-utils, resulting in a single request rather than many. This
    function will parse information from the files and then delete them. Returns
    a Dataset object, holding information about all the associated experiments
    and biosamples.
    '''
    if acc.startswith('GSE') or '/GSE' in acc:  # experiment series
        if '/' in acc:
            gse = GEOparse.get_GEO(filepath=acc)
        else:
            gse = GEOparse.get_GEO(geo=acc)  # pragma: no cover
        # create Experiment objects from each GSM file
        experiments = [obj for obj in [parse_gsm(gsm, experiment_type) for gsm in gse.gsms.values()] if
                       obj and obj.public]
        # delete file after GSMs are parsed
        if '/' not in acc:  # pragma: no cover
            print('GEO parsing done. Removing downloaded soft file.')
            os.remove(f'{acc}_family.soft.gz')
        if not experiments:
            print('Sequencing experiments not found. Exiting.')
            return
        gds = Dataset(acc, gse.metadata['sample_id'], experiments,
                      [parse_bs_record(experiment.bs) for experiment in experiments])
        return gds
    elif acc.startswith('GSM') or '/GSM' in acc:  # single experiment
        if '/' in acc:
            gsm = GEOparse.get_GEO(filepath=acc)
        else:
            gsm = GEOparse.get_GEO(geo=acc)  # pragma: no cover
        exp = parse_gsm(gsm, experiment_type)
        print("GEO parsing done. Removing downloaded soft file.")
        try:
            os.remove(f'{acc}.txt')  # delete file after GSM is parsed
        except Exception:
            pass
        if not exp:
            print("Accession not a sequencing experiment, or couldn't be parsed. Exiting.")
            return
        gds = Dataset(None, [acc], [exp], [parse_bs_record(exp.bs)])
        return gds
    else:
        print('Input not a valid GEO accession.')
        return


def parse_bs_record(bs_acc):
    '''
    Takes in a BioSample accession, fetches the BioSample record, and
    parses it into a Biosample object.
    '''
    print("Fetching BioSample record...")
    bs_handle = handle_timeout(Entrez.efetch(db='biosample', id=bs_acc))
    bs_xml = ET.fromstring(bs_handle.read())
    atts = {}
    descr = ''
    acc = bs_xml.find('./BioSample').attrib['accession']
    org = [item.text for item in bs_xml.iter("OrganismName")]
    # turn org items such as "Mus musculus x Mus spretus" to Mus musculus
    org = ['Mus musculus' if 'mus musculus' in item.lower() else item for item in org]
    if not org:
        org = [item.attrib['taxonomy_name'] for item in bs_xml.iter("Organism")]
    # except IndexError:
    #     print("{} - can't get OrganismName".format(bs_acc))
    for item in bs_xml.iter("Attribute"):
        atts[item.attrib['attribute_name']] = item.text

    for name in ['source_name', 'sample_name', 'gender', 'strain', 'genotype', 'cross',
                 'cell_line', 'cell line', 'cell lines', 'tissue', 'sirna transfected', 'treatment', 'activation time']:
        if name in atts.keys() and atts[name].lower() != 'none':
            if atts[name] not in descr:
                descr += name + ": " + atts[name] + ' | '
            if name == 'treatment':
                treatments = atts[name]
                if not sum([term in treatments.lower() for term in ['blank', 'none', 'n/a']]):
                    # print message to indicate that Treatment tab will need to be filled
                    print(f"BioSample accession {acc} has treatment attribute",
                          "but treatment not written to file")
    descr = descr.rstrip(' | ')
    bs = Biosample(acc, org[0], descr)
    return bs


def get_geo_tables(geo_acc, outf, lab_alias='4dn-dcic-lab', email='', types=type_dict):
    '''
    Creates 3 separate tsv files containing information for fastq files,
    experiments, and biosamples associated with a GEO accession. Can be used if a
    blank workbook with the required Experiment sheets hasn't been created.

    Parameters:
    geo_acc - GEO accession (e.g. 'GSE93431')
    lab_alias - alias prefix; default is '4dn-dcic-lab'
    outf - prefix for output files. Output files will be named
           <outf>_expts.tsv, <outf>_fqs.tsv, and <outf>_bs.tsv.
    email - email to be supplied for NCBI Entrez e-utils.
    types - dictionary of experiment types - do not use, leave default.
    '''
    Entrez.email = email if email else input('Enter email address to use NCBI Entrez: ')
    gds = get_geo_metadata(geo_acc, experiment_type=None)
    with open(outf + '_expts.tsv', 'w') as outfile:
        for exp in gds.experiments:
            outfile.write(f'{lab_alias}:{exp.geo}\t{exp.title}\t{types[exp.exptype]}\t{exp.bs,}\t{",".join(exp.runs)}\tGEO:{exp.geo}\n')
    with open(outf + '_fqs.tsv', 'w') as outfile:
        for exp in gds.experiments:
            if exp.layout == 'single':  # single end reads
                for run in exp.runs:
                    outfile.write(f'{lab_alias}:{run}_fq\t{exp.title}\tfastq\t \t \t \t{str(exp.length)}\t{exp.instr}\tSRA:{run}\n')
            elif exp.layout == 'paired':  # paired end reads
                for run in exp.runs:
                    alias = lab_alias + ':' + run
                    outfile.write(f'{alias}_fq1\t{exp.title}\tfastq\t1\tpaired with\t{alias}_fq2\t{str(exp.length)}\t{exp.instr}\tSRA:{run}\n')
                    outfile.write(f'{alias}_fq2\t{exp.title}\tfastq\t2\tpaired with\t{alias}_fq1\t{str(exp.length)}\t{exp.instr}\tSRA:{run}\n')
    with open(outf + '_bs.tsv', 'w') as outfile:
        for biosample in gds.biosamples:
            outfile.write(f'{lab_alias}:{biosample.acc}\t{biosample.description}\tBioSample:{biosample.acc}\n')


def append_xlsx_rows_unformatted(sheet, content_dict):
    '''
        Appends content as a row to the given xlsx sheet. The sheet must have
        column names in the first row. The content to append must be formatted
        as a dictionary with column:value where column matches a column name
        from the sheet and value is a string. Dictionary keys that do not match
        the sheet columns are ignored. Sheet columns for which no key is present
        in the dictionary will get an empty string. The row is appended at the
        bottom of the sheet.
    '''
    fields = list_column_titles(sheet)
    content_row = ['' for i in range(len(fields))]
    for field in content_dict:
        if field in fields:
            content_row[fields.index(field)] = content_dict[field]
    sheet.append(content_row)
    return sheet


def write_experiments(sheet_name, experiments, alias_prefix, file_dict, inbook, outbook, types=type_dict):
    '''
    Writes relevant Experiment object attributes to an Experiment sheet.
    Possible sheet types: ExperimentSeq, ExperimentHiC, ExperimentRepliseq,
                          ExperimentAtacseq, ExperimentDamid, ExperimentChiapet
    Writes alias, description, biosample, files, dbxrefs, and experiment_type
    fields, as appropriate.
    '''
    sheet = outbook[sheet_name]
    print(f"Writing {sheet_name} sheet...")
    for exp in experiments:
        content_dict = {
            'aliases': alias_prefix + ':' + exp.geo,
            'description': exp.title,
            'dbxrefs': 'GEO:' + exp.geo
        }
        if 'Biosample' in inbook.sheetnames:
            content_dict['*biosample'] = alias_prefix + ':' + exp.bs
        if 'FileFastq' in inbook.sheetnames:
            content_dict['files'] = ','.join(file_dict[exp.geo])
        if exp.link:
            content_dict['dbxrefs'] += ', SRA:' + exp.link
        if exp.exptype in types:
            content_dict['*experiment_type'] = types[exp.exptype]
        sheet = append_xlsx_rows_unformatted(sheet, content_dict)
    return outbook


def experiment_type_compare(sheetname, expt_list, geo, inbook):
    '''
    For a given experiment type, looks for that type in workbook sheets and compares
    to experiment types of GEO record/dataset. If present in both, will write
    experiments to file; if either is missing, will print an warning message.
    '''
    expt_dict = {'Atacseq': 'ATAC-seq', 'Damid': 'DamID', 'Chiapet': 'ChIA-PET',
                 'Seq': 'ChIP-seq, RNA-seq, SPRITE, or TSA-seq'}
    expt_name = sheetname[10:] if sheetname[10:] not in expt_dict.keys() else expt_dict[sheetname[10:]]
    type_name = sheetname[10:] if sheetname != 'ExperimentSeq' else '<experiment_type>'
    if sheetname in inbook.sheetnames and expt_list:
        return True
    elif sheetname in inbook.sheetnames and not expt_list:
        print(f"\nNo {expt_name} experiments parsed from {geo}.")
        print(f"If all samples are known to be {expt_name} experiments,")
        print(f"this script can be rerun using -t {type_name}")
    elif sheetname not in inbook.sheetnames and expt_list:
        print(f"\n{expt_name} experiments found in {geo} but no {sheetname} sheet")
        print("present in workbook.",
              expt_name if sheetname != 'ExperimentSeq' else 'These',
              "experiments will not be written to file.")
    return False


def modify_xlsx(geo, infile, outfile, alias_prefix, experiment_type=None, types=valid_types):
    '''
    Looks up a GEO Series record, parses it along with its associated SRA and
    BioSample records, and writes relevant attributes to the specified file. An
    excel template workbook must be specified, and for each type of metadata
    object, will look for the relevant sheet in the workbook. If sheet is absent
    these won't get written.
    '''
    gds = get_geo_metadata(geo, experiment_type)
    if not gds:
        return
    book = openpyxl.load_workbook(infile)

    # copy book to outbook
    outbook = openpyxl.Workbook()
    outbook.remove(outbook.active)  # removes the empty sheet created by default named Sheet
    for sheet_name in book.sheetnames:
        new_sheet = copy_xlsx_sheet(book[sheet_name], outbook, sheet_name)[0]

    get_organisms = requests.get('https://data.4dnucleome.org/search/?type=Organism&frame=object&format=json')
    organisms = [d['scientific_name'] for d in get_organisms.json()['@graph'] if d.get('scientific_name')]
    bs_to_write = [bs.acc for bs in gds.biosamples if bs.organism in organisms]
    exp_to_write = [exp for exp in gds.experiments if exp.public and exp.bs in bs_to_write]

    exp_sheets = [name for name in book.sheetnames if name.startswith('Experiment')]
    if len(exp_sheets) > 0:
        # looks for each experiment type in parsed data
        # then looks for relevant worksheet in excel template
        # writes experiments to file if both present
        hic_expts = [exp for exp in exp_to_write if exp.exptype.startswith('hic')
                     or exp.exptype.startswith('dnase hic')]
        seq_expts = [exp for exp in exp_to_write if exp.exptype in
                     ['chipseq', 'rnaseq', 'tsaseq'] or 'sprite' in exp.exptype]
        atac_expts = [exp for exp in exp_to_write if exp.exptype == 'atacseq']
        rep_expts = [exp for exp in exp_to_write if exp.exptype == 'repliseq']
        dam_expts = [exp for exp in exp_to_write if exp.exptype.startswith('damid')]
        cap_expts = [exp for exp in exp_to_write if exp.exptype == 'capturec']
        chia_expts = [exp for exp in exp_to_write if exp.exptype in ['chiapet', 'placseq']]
        sheet_types = {'HiC': hic_expts, 'Seq': seq_expts, 'Damid': dam_expts,
                       'Atacseq': atac_expts, 'Repliseq': rep_expts,
                       'CaptureC': cap_expts, 'Chiapet': chia_expts}

        keep = []
        keep_keys = []
        for key in sheet_types.keys():
            if experiment_type_compare('Experiment' + key, sheet_types[key], geo, book):
                keep += sheet_types[key]
                keep_keys.append(key)

    if 'Biosample' in book.sheetnames:
        bs_sheet = outbook['Biosample']
        print("Writing Biosample sheet...")
        for entry in gds.biosamples:
            if entry.acc in bs_to_write and entry.acc in [item.bs for item in keep]:
                # write each Biosample object to file
                content_dict = {
                    'aliases': alias_prefix + ':' + entry.acc,
                    'description': entry.description,
                    'dbxrefs': 'BioSample:' + entry.acc
                }
                if 'BiosampleCellCulture' in book.sheetnames:
                    content_dict['cell_culture_details'] = alias_prefix + ':' + entry.acc + '-cellculture'
                bs_sheet = append_xlsx_rows_unformatted(bs_sheet, content_dict)

    if 'BiosampleCellCulture' in book.sheetnames:
        bcc_sheet = outbook['BiosampleCellCulture']
        print("Writing BiosampleCellCulture sheet...")
        for entry in gds.biosamples:
            if entry.acc in bs_to_write and entry.acc in [item.bs for item in keep]:
                # generate aliases for BiosampleCellCulture sheet
                content_dict = {
                    'aliases': alias_prefix + ':' + entry.acc + '-cellculture'
                }
                bcc_sheet = append_xlsx_rows_unformatted(bcc_sheet, content_dict)

    file_dict = {}
    if 'FileFastq' in book.sheetnames:
        fq_sheet = outbook['FileFastq']
        print("Writing FileFastq sheet...")
        for entry in gds.experiments:
            if entry.bs in bs_to_write and entry in keep:
                file_dict[entry.geo] = []
                for run in entry.runs:
                    # write information about SRA runs to file -
                    # assumes they will be downloaded as fastq files
                    content_dict = {
                        '*file_format': 'fastq',
                        'read_length': entry.length,
                        'instrument': entry.instr,
                        'dbxrefs': 'SRA:' + run
                    }
                    if entry.layout.lower() == 'paired':
                        fq1 = alias_prefix + ':' + run + '_1_fq'
                        fq2 = alias_prefix + ':' + run + '_2_fq'
                        file_dict[entry.geo] += [fq1, fq2]
                        content_dict_fq1 = {k: v for k, v in content_dict.items()}
                        content_dict_fq1['aliases'] = fq1
                        content_dict_fq1['paired_end'] = '1'
                        content_dict_fq1['related_files.relationship_type'] = 'paired with'
                        content_dict_fq1['related_files.file'] = fq2
                        fq_sheet = append_xlsx_rows_unformatted(fq_sheet, content_dict_fq1)
                        content_dict_fq2 = {k: v for k, v in content_dict.items()}
                        content_dict_fq2['aliases'] = fq2
                        content_dict_fq2['paired_end'] = '2'
                        fq_sheet = append_xlsx_rows_unformatted(fq_sheet, content_dict_fq2)

                    elif entry.layout.lower() == 'single':
                        fq_0 = alias_prefix + ':' + run + '_fq'
                        file_dict[entry.geo] += [fq_0]
                        content_dict['aliases'] = fq_0
                        fq_sheet = append_xlsx_rows_unformatted(fq_sheet, content_dict)

                    else:
                        raise ValueError("Invalid value for layout. Layout must be 'single' or 'paired'.")

    if len(exp_sheets) > 0:
        for key in keep_keys:
            outbook = write_experiments('Experiment' + key, sheet_types[key],
                                        alias_prefix, file_dict, book, outbook)

        other_organisms = [exp.geo for exp in gds.experiments if exp.bs not in bs_to_write]
        if other_organisms:
            print('\nThe following accessions were from non-4DN organisms and were not written to file:')
            print('\n'.join(other_organisms))
        other = [exp for exp in gds.experiments if exp.exptype not in types]
        skip = ['bisulfiteseq', 'groseq', 'smartseq', '4cseq']
        skipped = [e for e in other if e.exptype in skip]
        if skipped:
            print('\nThe following accessions had non-4DN experiment types and were not written to file:')
            print('\n'.join([item.geo for item in skipped]))
            other = [e for e in other if e.exptype not in skip]
        if other:
            if len(other) + len(skipped) == len(gds.experiments):
                print(f"\nExperiment types of dataset could not be parsed. {', '.join(exp_sheets)} sheet not written")
            else:
                print("\nThe following accessions had experiment types that could not be parsed:")
                for item in other:
                    print(item.geo)
            print("If these samples are of a single known experiment type,",
                  "this script can be rerun using -t <experiment_type>")

    outbook.save(outfile)
    print(f"\nWrote file to {outfile}.")
    return


def main(types=valid_types, descr=description, epilog=epilog):  # pragma: no cover
    parser = argparse.ArgumentParser(description=descr, epilog=epilog,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument('geo_accession', help="GEO accession", action="store")
    parser.add_argument('-i', '--infile', help="Input xlsx file - blank submit4dn workbook",
                        action="store", required=True)
    parser.add_argument('-o', '--outfile', help="Output xlsx file - default output \
                        filename will be GEO accession with xlsx extension",
                        default='', action="store")
    parser.add_argument('-a', '--alias', help="Alias prefix, default is '4dn-dcic-lab'",
                        action="store", default="4dn-dcic-lab")
    parser.add_argument('-e', '--email', help="Email address to use NCBI Entrez",
                        action="store", default="")
    parser.add_argument('-k', '--apikey', help="API Key for NCBI Entrez",
                        action="store", default="")
    parser.add_argument('-t', '--type', help="Optional: type of experiment in series. \
                        By default experiment type is parsed from SRA records, but \
                        this option is useful when parsing isn't straightforward. \
                        Accepted types: HiC, ChIP-seq, RNA-seq, TSA-seq, ATAC-seq, DamID, Repliseq, \
                        DNase HiC, Capture HiC, PLAC-seq, DNA SPRITE, RNADNA SPRITE. \
                        Note that only one type may be specified, so make sure GEO Series \
                        doesn't include multiple experiment types.",
                        action="store", default=None)
    args = parser.parse_args()
    out_file = args.outfile if args.outfile else args.geo_accession + '.xlsx'
    if args.type and args.type.lower().replace('-', '') not in types:
        print(f"\nError: {args.type} not a recognized type\n")
        parser.print_help()
        sys.exit()
    if not args.email:
        Entrez.email = input('Enter email address to use NCBI Entrez: ')
    else:
        Entrez.email = args.email
    if args.apikey:
        Entrez.api_key = args.apikey
    modify_xlsx(args.geo_accession, args.infile, out_file, args.alias, experiment_type=args.type)


if __name__ == '__main__':  # pragma: no cover
    main()
